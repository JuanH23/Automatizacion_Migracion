import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import Color,Font
import openpyxl
import os
from dotenv import set_key,dotenv_values 
from Advertencia import*
import numpy as np
import random
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtCore import QTimer

def diseño(df,df_cos_daas,name_file,filter_daas,type_node): # Crear DataFrame con información

    ###########
    env=dotenv_values(".env")
    ruth_list_download= env["path_list_download"]
    ruta_nueva_carpeta = ruth_list_download + "/Diseños_NODOS"
    file_name="" 
    os.makedirs(ruta_nueva_carpeta, exist_ok=True)
    if type_node == "1 x 2":
        file_name='Diseño Segmentaciones RPHY-NODO '   
    elif type_node =="2 x 4":
        file_name='Diseño Segmentaciones RPHY 2x4-NODO '

    ruta_archivo = os.path.join(ruta_nueva_carpeta,file_name+ name_file +'.xlsx')

    ###########
    archivo_excel = pd.ExcelWriter(ruta_archivo, engine='openpyxl')     # Crear archivo Excel desde cero y escribir información del DataFrame

    df_cd=pd.DataFrame(df_cos_daas)                                     # Escribir el DataFrame en el archivo Excel
    print(f"df_cd==>{df_cd}")
    ################################################################ 
    df.to_excel(archivo_excel,sheet_name='Hoja1' ,index=False)
    ################################################################
    hoja = archivo_excel.sheets['Hoja1']
    
    workbook = archivo_excel.book   # Obtener el libro de trabajo y la hoja
    worksheet = workbook.active
    
    #######################
    # Escribir la primera columna del dataframe en la columna A de la hoja de trabajo
    columna = 1  # Columna A
    fila_inicial = 4  # Empezar a escribir desde la fila 4
    for i, valor in enumerate(df['CMTS']):
        celda_actual = worksheet.cell(row=fila_inicial+i, column=columna)
        celda_actual.value = valor
    # Escribir la primera columna del dataframe en la columna B de la hoja de trabajo
    columna = 2  # Columna B
    fila_inicial = 4  # Empezar a escribir desde la fila 4
    for i, valor in enumerate(df['Up']):
        celda_actual = worksheet.cell(row=fila_inicial+i, column=columna)
        celda_actual.value = valor
    # Escribir la primera columna del dataframe en la columna E de la hoja de trabajo
    columna = 5  # Columna E
    fila_inicial = 4  # Empezar a escribir desde la fila 4
    
    for i, valor in enumerate(df['Total'].astype(int)):
        celda_actual = worksheet.cell(row=fila_inicial+i, column=columna)
        celda_actual.value = valor
    # Escribir la primera columna del dataframe en la columna F de la hoja de trabajo
    columna = 6  # Columna F
    fila_inicial = 4  # Empezar a escribir desde la fila 4
    for i, valor in enumerate(df['Description']):
        celda_actual = worksheet.cell(row=fila_inicial+i, column=columna)
        celda_actual.value = valor
    
    columna = worksheet["C"]
    rango=columna[3:12]
    
    for celda in rango:     # bucle para establecer el valor de cada celda en None
        celda.value = None
        
    columna = worksheet["D"]
    rango=columna[3:12]

    for celda in rango:     # bucle para establecer el valor de cada celda en None
        celda.value = None             
    #######################


    
    fill = PatternFill(fill_type='solid', start_color='FFFFFFFF', end_color='FFFFFFFF')     # Crear objeto de relleno blanco

    
    for sheet_name in workbook.sheetnames:      # Recorrer todas las celdas y aplicar el relleno blanco
        sheet = workbook[sheet_name]
        

        # Obtener el número de filas y columnas de la hoja
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        for row in range(1, max_row + 1):       # Recorrer todas las celdas de la hoja y establecer su color de fondo en blanco
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.fill = fill

    ########################################TABLA_ANTES##########################################
    #Unir celdas en especifico
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    hoja.merge_cells(start_row=4, start_column=7, end_row=7, end_column=7)
    #Se le da un tamaño que se requiere a la fila y a la columna "A" y "F"
    hoja.row_dimensions[1].height=20
    hoja.column_dimensions['A'].width=28
    hoja.column_dimensions['F'].width=40
    #Tamaño de la letra
    fuente=Font(size=14)
    fuente.bold=True
    #Aplica los cambios dados 
    hoja.cell(row=1,column=1).font=fuente
    #obtiene el nombre del CMTS para colocar el texto en la celda
    chasis_valor=df['CMTS']
    chasis_index=chasis_valor.index
    chasis_list=chasis_index.to_list()
    print(f"chasis_valor==>{chasis_valor}")
    print(f"chasis_index==>{chasis_list}")
    indice=chasis_list[1]
    texto=df.loc[indice,"CMTS"]
    #texto_chasis="MEDE-CABA-H-03-CS100G#"
    celda=hoja.cell(row=1,column=1)
    celda.value=texto
    hoja.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    hoja.row_dimensions[2].height=30
    hoja.column_dimensions['F'].width=40
    fuente=Font(size=12)
    fuente.bold=True
    hoja.cell(row=2,column=1).font=fuente
    texto="Antes"
    celda=hoja.cell(row=2,column=1)
    celda.value=texto
    cell = worksheet.cell(row=2, column=1)
    cell_range = worksheet['A2:G2']
    #Centra el texto en la celda tanto horizontal como verticalmente
    cell.alignment = Alignment(horizontal='center', vertical='center')
    #Le da un tipo de relleno al borde de las celdas en este caso  "thick"
    border=Border(top=Side(style='thick'),bottom=Side(style='thick'),left=Side(style='thick'),right=Side(style='thick'))
    border_chasis=Border(top=Side(style='thick'),bottom=Side(style='thick'),left=Side(style='thick'),right=Side(style='thick'))
    #aplica los cambios en el rango de celda dado
    for cells in cell_range:
        for cell in cells:
            cell.border = border
    #Texto de las cabeceras de la tabla
    texto="CHASIS"
    celda_CHASIS=hoja.cell(row=3,column=1)
    celda_CHASIS.value=texto
    texto="SLOT"
    celda=hoja.cell(row=3,column=2)
    celda.value=texto
    texto="CMAC"
    celda=hoja.cell(row=3,column=3)
    celda.value=texto
    texto="DMAC"
    celda=hoja.cell(row=3,column=4)
    celda.value=texto
    texto="CLIENTES"
    celda=hoja.cell(row=3,column=5)
    celda.value=texto
    texto="NOMBRE"
    celda=hoja.cell(row=3,column=6)
    celda.value=texto
    texto="USUARIOS"
    celda_USU=hoja.cell(row=3,column=7)
    celda_USU.value=texto
    celda_USU.border=border_chasis
    fuente=Font(size=10)
    fuente.bold=True
    hoja.cell(row=3,column=7).font=fuente
    cell_aligment_row3= Alignment(horizontal='center', vertical='center')
    fila = 3 # Aquí se selecciona la fila deseada
    for celda in hoja[fila]:
        celda.alignment = cell_aligment_row3
    hoja.column_dimensions['G'].width=10
    #"cell_range_row3": Rango de celda que se van a aplicar los cambios 
    cell_range_row3 = worksheet['A3:F3']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    #Crear objeto de color azul
    dark_blue=Color(rgb='366092')
    relleno = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type='solid')
    border=Border(top=Side(style='thin'),bottom=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'))
    white_font = Font(color='FFFFFF')
    for cells in cell_range_row3:
        for cell in cells:
            cell.border = border
            cell.font=white_font
            cell.fill=relleno

    rango=hoja['E4:E12']
    
    #Suma los valores que tenga en el rango que sea dependiendo del nodo
    suma_columna = sum([0 if celda.value is None else celda.value for fila in rango for celda in fila])

    celda_resultado = hoja['G4']
    celda_resultado.value = suma_columna
    cell = worksheet.cell(row=4, column=7)
    red=Color(rgb='FF0000')
    red_font = Font(color=red)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font=red_font
    ########################################TABLA_DEPUES##########################################
    #combina dos celdas y da tamaño en horizontal por columna
    hoja.merge_cells(start_row=1, start_column=9, end_row=1, end_column=17)
    hoja.column_dimensions['I'].width=20
    hoja.column_dimensions['J'].width=25
    hoja.column_dimensions['K'].width=15
    hoja.column_dimensions['L'].width=25
    hoja.column_dimensions['N'].width=15
    hoja.column_dimensions['Q'].width=45
    fuente=Font(size=14)
    fuente.bold=True
    hoja.cell(row=1,column=9).font=fuente
    Dispositivo_cos_valor=df_cos_daas['Dispositivo COS']
    Dispositivo_cos_index=Dispositivo_cos_valor.index
    Dispositivo_cos_list=Dispositivo_cos_index.to_list()
    indice_Dispositivo_cos=Dispositivo_cos_list[1]
    texto_dispositivo_cos=df_cos_daas.loc[indice_Dispositivo_cos,"Dispositivo COS"]
    #texto_dispositivo_cos="BOGO-FONT-H-09-COS"
    celda=hoja.cell(row=1,column=9)
    celda.value=texto_dispositivo_cos
    hoja.merge_cells(start_row=2, start_column=9, end_row=2, end_column=17)
    hoja.merge_cells(start_row=4, start_column=9, end_row=8, end_column=9)
    fuente=Font(size=12)
    fuente.bold=True
    hoja.cell(row=2,column=9).font=fuente
    texto="DESPUES"
    celda=hoja.cell(row=2,column=9)
    celda.value=texto
    cell1 = worksheet.cell(row=1, column=9)
    cell2 = worksheet.cell(row=2, column=9)
    #Delimita el rango a donde se van a aplicar los cambios
    cell_range_row4 = worksheet['I2:Q2']
    cell_range_row5 = worksheet['I3:Q3']
    cell_range_row6 = worksheet['I4:Q4']
    cell_range_row7 = worksheet['I5:Q5']
    cell_range_row8 = worksheet['I7:Q7']
    cell_range_row9 = worksheet['I8:Q8']
    cell_range_row10 = worksheet['I6']
    cell1.alignment = Alignment(horizontal='center', vertical='center')
    cell2.alignment = Alignment(horizontal='center', vertical='center')
    border=Border(top=Side(style='thin'),bottom=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'))
    cell_range_row10.border=border
    #Aplica los cambios en el borde a cada una de las celdas

    for cells in cell_range_row4:
        for cell in cells:
            cell.border = border
    for cells in cell_range_row5:
        for cell in cells:
            cell.border = border
    for cells in cell_range_row6:
        for cell in cells:
            cell.border = border
    for cells in cell_range_row7:
        for cell in cells:
            cell.border = border
    for cells in cell_range_row8:
        for cell in cells:
            cell.border = border
    for cells in cell_range_row9:
        for cell in cells:
            cell.border = border      
    #Asigna el texto a las cabeceras de la tabla del despues                   
    texto="REGIONAL"
    celda=hoja.cell(row=3,column=9)
    celda.value=texto
    texto="DAAS"
    celda=hoja.cell(row=3,column=10)
    celda.value=texto
    texto="PUERTO DAAS"
    celda=hoja.cell(row=3,column=11)
    celda.value=texto
    texto="CHASIS"
    celda=hoja.cell(row=3,column=12)
    celda.value=texto
    texto="RPD"
    celda=hoja.cell(row=3,column=13)
    celda.value=texto
    texto="UPSTREAM"
    celda=hoja.cell(row=3,column=14)
    celda.value=texto
    texto="DMAC"
    celda=hoja.cell(row=3,column=15)
    celda.value=texto     
    texto="CLIENTES"
    celda=hoja.cell(row=3,column=16)
    celda.value=texto 
    texto="NOMBRE"
    celda=hoja.cell(row=3,column=17)
    celda.value=texto 
    cell_range_row3_despues= worksheet['I3:Q3']
    alignment = Alignment(horizontal='center', vertical='center')
    border=Border(top=Side(style='thin'),bottom=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin'))
    white_font = Font(color='FFFFFF')
    dark_blue=Color(rgb='366092')
    relleno = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type='solid')
    for cells in cell_range_row3_despues:
        for cell in cells:
            cell.border = border
            cell.font=white_font
            cell.fill=relleno
            cell.alignment=alignment

    cell = worksheet.cell(row=4, column=9)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    yellow=Color(rgb='FFC000')
    relleno_yellow = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
    pink=Color(rgb='FCD5B4')
    relleno_pink=PatternFill(start_color=pink,end_color=pink,fill_type='solid')
    #Delimita el rango a donde se van a aplicar los cambios
    cell_range_row4=worksheet['J4:Q4']
    cell_range_row5=worksheet['J5:Q5']
    cell_range_row7=worksheet['J7:Q7']
    cell_range_row8=worksheet['J8:Q8']
    #Aplica los cambios en el borde a cada una de las celdas
    for cells in cell_range_row4:
        for cell in cells:
            cell.fill=relleno_yellow 
    for cells in cell_range_row5:
        for cell in cells:
            cell.fill=relleno_pink         
    for cells in cell_range_row7:
        for cell in cells:
            cell.fill=relleno_yellow 
    for cells in cell_range_row8:
        for cell in cells:
            cell.fill=relleno_pink

    ###############################################
    DISPOSITIVO_DAAS = df_cd['Dispositivo DAAS'].unique()
    simil=[]
    #Realiza la conversion del segundo das en vez que comience de 0-48, empiece
    #de 49-97.
    if df_cd['Dispositivo DAAS'].str.contains(str(filter_daas+1)).any():#Revisa que tenga mas de un DAAS
            print("ENTRO AL DAAS")
            print(f"filter_DAAS==>{filter_daas+1}")
            #Organiza los valores 
            df_cd = df_cd.sort_values('Puerto COS',inplace=False,ascending=True)
            #Crea una mascara
            mask_range = df_cd['Puerto DAAS'].between('xe-0/0/0', 'xe-0/0/48')
            mask_name = df_cd['Dispositivo DAAS'].str.contains(str(filter_daas+1))
            mask_range_name = mask_name & mask_range
            #Le suma a cada valor que ya se le realizo un split con el simbolo "/" del segundo Daas y le suma 49 a cada valor que encuentre
            df_cd.loc[mask_range_name, 'Puerto DAAS'] = (
                df_cd.loc[mask_range_name, 'Puerto DAAS']
                .str.replace(r'xe-0/0/(\d+)', lambda x: 'xe-0/0/' + str(int(x.group(1))+49))
            )
            
            df_cd['Puerto DAAS']=df_cd['Puerto DAAS'].astype(str)
            df_cd['ultimo_num_DAAS'] = df_cd['Puerto DAAS'].apply(lambda x: get_x(x, 0))
            # Extraer el primer número de cada entrada en la columna puerto_COS
            df_cd['primer_num_COS'] = df_cd['Puerto COS'].str.split(':').str[0]
            print(f"df_cd==>{df_cd}")
            df_cd=df_cd.drop_duplicates(subset='primer_num_COS')
            print(f"df_cd_sin_duplicados==>{df_cd}")
            #df_cd.to_excel("new_numbers.xlsx")
            ###############################!
            #Crea una columna con los numeros unicos del DAAS
            numeros_coincidentes=df_cd['ultimo_num_DAAS'].unique()
            #Revisa cuales de los puertos del DAAS estan en la columna nueva donde se encuentran
            coincidente_COS=df_cd[df_cd['primer_num_COS'].isin(numeros_coincidentes)]
            #coincidente=df_cd.loc[df_cd['ultimo_num_DAAS'].isin(df_cd['primer_num_COS'])]
            #coincidente_COS.to_excel("coincidente.xlsx")
            coincidente_DAAS=df_cd[df_cd['ultimo_num_DAAS'].isin(coincidente_COS['primer_num_COS'])]
            #coincidente_DAAS.to_excel("coincidente_2.xlsx")
            coincidente_COS=coincidente_COS.loc[:,['Dispositivo COS','Puerto COS','primer_num_COS']]
            coincidente_COS=coincidente_COS.reset_index(drop=True)
            coincidente_DAAS=coincidente_DAAS.loc[:,['Dispositivo DAAS','Puerto DAAS']]
            coincidente_DAAS=coincidente_DAAS.reset_index(drop=True)
            #une ambos Dataframe para trabajar con un solo Data
            merge_coincidente=pd.concat([coincidente_COS,coincidente_DAAS],axis=1)
            merge_coincidente.to_excel('merge_coincidente.xlsx')

            valores_unicos=merge_coincidente['primer_num_COS'].unique().tolist()
            #Toma un valor aleatorio del puerto COS y lo coloca en el formato
            valor_aleatorio = random.choice(valores_unicos)
            valores_unicos.remove(valor_aleatorio)
            filas_aleatorias = merge_coincidente.loc[merge_coincidente['primer_num_COS'] == valor_aleatorio]
            
            print(f"numero_random_solo_UNA__VEZ==>{filas_aleatorias}")
            print(f"TYPE_NODE==>{type_node}")
            ###############################!
              
            #print(f"simil==>{simil}")
    else:#Si solo tiene un dispositivo DAAS
            df_cd['Puerto DAAS']=df_cd['Puerto DAAS'].astype(str)
            df_cd['ultimo_num_DAAS'] = df_cd['Puerto DAAS'].apply(lambda x: get_x(x, 0))

            # Extraer el primer número de cada entrada en la columna puerto_COS
            df_cd['primer_num_COS'] = df_cd['Puerto COS'].str.split(':').str[0]
            df_cd=df_cd.drop_duplicates(subset='primer_num_COS')
            #df_cd.to_excel("same_new_numbers.xlsx")
            ###############################!
            numeros_coincidentes=df_cd['ultimo_num_DAAS'].unique()
            coincidente_COS=df_cd[df_cd['primer_num_COS'].isin(numeros_coincidentes)]
            #coincidente=df_cd.loc[df_cd['ultimo_num_DAAS'].isin(df_cd['primer_num_COS'])]
            #coincidente_COS.to_excel("coincidente.xlsx")
            coincidente_DAAS=df_cd[df_cd['ultimo_num_DAAS'].isin(coincidente_COS['primer_num_COS'])]
            #coincidente_DAAS.to_excel("coincidente_2.xlsx")
            coincidente_COS=coincidente_COS.loc[:,['Dispositivo COS','Puerto COS','primer_num_COS']]
            coincidente_COS=coincidente_COS.reset_index(drop=True)
            coincidente_DAAS=coincidente_DAAS.loc[:,['Dispositivo DAAS','Puerto DAAS']]
            coincidente_DAAS=coincidente_DAAS.reset_index(drop=True)
            merge_coincidente=pd.concat([coincidente_COS,coincidente_DAAS],axis=1)
            #merge_coincidente.to_excel('merge_coincidente.xlsx')
            
            print(f"TYPE_NODE==>{type_node}")
            ###############################!    
            merge_coincidente.to_excel("merge_same_numbers.xlsx")  
            valores_unicos=merge_coincidente['primer_num_COS'].unique().tolist()
            valor_aleatorio = random.choice(valores_unicos)
            valores_unicos.remove(valor_aleatorio)
            filas_aleatorias = merge_coincidente.loc[merge_coincidente['primer_num_COS'] == valor_aleatorio]
            filas_aleatorias['primer_num_COS']=filas_aleatorias['primer_num_COS'].astype(str)
            print(f"numero_random_solo_UNA__VEZ==>{filas_aleatorias}")
            
            print(f"TYPE_NODE==>{type_node}")       
            print(f"simil==>{simil}")
    ######################################!
    #Genera un tipo de formato u otro dependiendo si se coloco un nodo u otro
    if type_node == "1 x 2":
        texto="Nodo 1x2"
        celda=hoja.cell(row=4,column=9)
        celda.value=texto
    ################################!
        #Extrae de la columna 'Dispositivo DAAS',el texto para colocarlo en la celda
        slot_valor_DAAS=filas_aleatorias['Dispositivo DAAS']
        slot_index_DAAS=slot_valor_DAAS.index
        slot_list_DAAS=slot_index_DAAS.to_list()
        indice_slot_DAAS=slot_list_DAAS[0]
        texto_DAAS=filas_aleatorias.loc[indice_slot_DAAS,"Dispositivo DAAS"]    
        celda=hoja.cell(row=4,column=10)
        celda.value=texto_DAAS
        celda=hoja.cell(row=7,column=10)
        celda.value=texto_DAAS
        #Extrae de la columna 'primer_num_COS',el texto para colocarlo en la celda
        slot_valor_PUERTO_DAAS=filas_aleatorias['primer_num_COS']
        slot_index_PUERTO_DAAS=slot_valor_PUERTO_DAAS.index
        slot_list_PUERTO_DAAS=slot_index_PUERTO_DAAS.to_list()
        indice_slot_PUERTO_DAAS=slot_list_PUERTO_DAAS[0]
        text_num_generic=filas_aleatorias.loc[indice_slot_PUERTO_DAAS,"primer_num_COS"]  
        
        texto_Puerto_DAAS=text_num_generic
        texto_Puerto_DAAS=str(texto_Puerto_DAAS)
        
        celda=hoja.cell(row=4,column=11)
        celda.value=texto_Puerto_DAAS+"/0"
        celda=hoja.cell(row=7,column=11)
        celda.value=texto_Puerto_DAAS+"/0"
        #Extrae de la columna 'Dispositivo COS',el texto para colocarlo en la celda
        slot_valor_CHASIS=filas_aleatorias['Dispositivo COS']
        slot_index_CHASIS=slot_valor_CHASIS.index
        slot_list_CHAIS=slot_index_CHASIS.to_list()
        indice_slot_CHASIS=slot_list_CHAIS[0]
        texto_CHASIS=filas_aleatorias.loc[indice_slot_CHASIS,"Dispositivo COS"]
        celda=hoja.cell(row=4,column=12)
        celda.value=texto_CHASIS
        celda=hoja.cell(row=7,column=12)
        celda.value=texto_CHASIS
        texto_RPD=text_num_generic+":00"
        celda=hoja.cell(row=4,column=13)
        celda.value=texto_RPD
        celda=hoja.cell(row=7,column=13)
        celda.value=texto_RPD
        #texto_UPSTREAM=":0/0"
        #texto_UPSTREAM2=":0/1"        
        texto_UPSTREAM=text_num_generic+":0/0"
        texto_UPSTREAM2=text_num_generic+":0/1"
        celda=hoja.cell(row=4,column=14)
        celda.value=texto_UPSTREAM
        celda=hoja.cell(row=7,column=14)
        celda.value=texto_UPSTREAM2
        texto_DMAC=text_num_generic+":0/0.0"
        celda=hoja.cell(row=4,column=15)
        celda.value=texto_DMAC
        celda=hoja.cell(row=7,column=15)
        celda.value=texto_DMAC
        #Extrae de la columna 'Description',el texto para colocarlo en la celda
        nodo_valor=df['Description']
        nodo_index=nodo_valor.index
        nodo_list=nodo_index.to_list()
        indice_nodo=nodo_list[1]
        texto_NOMBRE=df.loc[indice_nodo,"Description"]
        tex=str(texto_NOMBRE)
        #busca los caracteres "(" y ")" para colocar el texto necesario, en este caso "3F"
        indice_find_1=tex.find("(")
        indice_find_2=tex.find(")")
        tex=tex[:indice_find_1]+"3F"+tex[indice_find_1:indice_find_2]+"3F"+tex[indice_find_2]
        #texto_NOMBRE=""
        celda=hoja.cell(row=4,column=17)
        celda.value=str(texto_NOMBRE)
        #celda.value=""
        celda=hoja.cell(row=7,column=17)
        celda.value=tex
        #celda.value=""
    ################################!
    #Si se selecciona nodo tipo "2 x 4"
    elif type_node == "2 x 4":
    ######################################!
        texto="Nodo 2x4"
        celda=hoja.cell(row=4,column=9)
        celda.value=texto
    ################################!
        #Extrae de la columna 'primer_num_COS',el texto para colocarlo en las celdas
        slot_valor_PUERTO_DAAS=filas_aleatorias['primer_num_COS']
        slot_index_PUERTO_DAAS=slot_valor_PUERTO_DAAS.index
        slot_list_PUERTO_DAAS=slot_index_PUERTO_DAAS.to_list()
        indice_slot_PUERTO_DAAS=slot_list_PUERTO_DAAS[0]
        text_num_generic=filas_aleatorias.loc[indice_slot_PUERTO_DAAS,"primer_num_COS"] 
        
        #Extrae de la columna 'Dispositivo DAAS',el texto para colocarlo en las celdas
        slot_valor_DAAS=filas_aleatorias['Dispositivo DAAS']
        slot_index_DAAS=slot_valor_DAAS.index
        slot_list_DAAS=slot_index_DAAS.to_list()
        indice_slot_DAAS=slot_list_DAAS[0]
        texto_DAAS=filas_aleatorias.loc[indice_slot_DAAS,"Dispositivo DAAS"]          
        #texto_DAAS=""    
        celda=hoja.cell(row=4,column=10)
        celda.value=texto_DAAS
        celda=hoja.cell(row=5,column=10)
        celda.value=texto_DAAS
        celda=hoja.cell(row=7,column=10)
        celda.value=texto_DAAS
        celda=hoja.cell(row=8,column=10)
        celda.value=texto_DAAS
        texto_Puerto_DAAS="3"
        texto_Puerto_DAAS_2=str(text_num_generic)
        #texto_Puerto_DAAS=str(text_num_generic)+"/0"
        celda=hoja.cell(row=4,column=11)
        celda.value=texto_Puerto_DAAS_2 +"/0"
        celda=hoja.cell(row=5,column=11)
        celda.value=texto_Puerto_DAAS_2 +"/0"
        celda=hoja.cell(row=7,column=11)
        celda.value=texto_Puerto_DAAS_2 +"/0"
        celda=hoja.cell(row=8,column=11)
        celda.value=texto_Puerto_DAAS_2 +"/0"
        #Extrae de la columna 'Dispositivo COS',el texto para colocarlo en las celdas
        slot_valor_CHASIS=filas_aleatorias['Dispositivo COS']
        slot_index_CHASIS=slot_valor_CHASIS.index
        slot_list_CHAIS=slot_index_CHASIS.to_list()
        indice_slot_CHASIS=slot_list_CHAIS[0]
        texto_CHASIS=filas_aleatorias.loc[indice_slot_CHASIS,"Dispositivo COS"]        
        #texto_CHASIS=""
        celda=hoja.cell(row=4,column=12)
        celda.value=texto_CHASIS
        celda=hoja.cell(row=5,column=12)
        celda.value=texto_CHASIS
        celda=hoja.cell(row=7,column=12)
        celda.value=texto_CHASIS
        celda=hoja.cell(row=8,column=12)
        celda.value=texto_CHASIS
        texto_RPD=":0"
        texto_RPD2=":1"
        texto_RPD_1=str(text_num_generic)+":0"
        texto_RPD_2=str(text_num_generic)
        #texto_RPD2=str(text_num_generic)+":1"
        celda=hoja.cell(row=4,column=13)
        celda.value=texto_RPD_1
        celda=hoja.cell(row=5,column=13)
        celda.value=texto_RPD_1
        celda=hoja.cell(row=7,column=13)
        celda.value=texto_RPD_2 + ":1"
        celda=hoja.cell(row=8,column=13)
        celda.value=str(text_num_generic) +":1"
        texto_UPSTREAM=str(text_num_generic) + ":0/0"
        texto_UPSTREAM_2=str(text_num_generic) + ":0/1"
        texto_UPSTREAM3=str(text_num_generic) + ":1/0"
        texto_UPSTREAM4=str(text_num_generic) + ":1/1"
        celda=hoja.cell(row=4,column=14)
        celda.value=texto_UPSTREAM
        celda=hoja.cell(row=5,column=14)
        celda.value=texto_UPSTREAM_2
        celda=hoja.cell(row=7,column=14)
        celda.value=texto_UPSTREAM3
        celda=hoja.cell(row=8,column=14)
        celda.value=texto_UPSTREAM4
        texto_DMAC_2=str(text_num_generic)
        texto_DMAC=":0/0.0"
        celda=hoja.cell(row=4,column=15)
        celda.value=texto_DMAC_2 + ":0/0.0"
        celda=hoja.cell(row=5,column=15)
        celda.value=texto_DMAC_2 + ":0/0.0"
        celda=hoja.cell(row=7,column=15)
        celda.value=texto_DMAC_2 + ":1/0.0"
        celda=hoja.cell(row=8,column=15)
        celda.value=texto_DMAC_2 + ":1/0.0"
        #Extrae de la columna 'Description',el texto para colocarlo en las celdas
        nodo_valor=df['Description']
        nodo_index=nodo_valor.index
        nodo_list=nodo_index.to_list()
        indice_nodo=nodo_list[1]
        text_NOMBRE=df.loc[indice_nodo,"Description"]
        texto_NOMBRE_2=str(text_NOMBRE)
        texto_NOMBRE="NODO"
        #busca los caracteres "(" y ")" para colocar el texto necesario, en este caso "2F","3F" y "4F", en cada una de las celdas correspondientes
        #ademas de asignar los valores 
        texx=str(texto_NOMBRE_2)
        indice_find__1=texx.find("(")
        indice_find__2=texx.find(")")
        tex_2=texx[:indice_find__1]+"2F"+texx[indice_find__1:indice_find__2]+"2F"+texx[indice_find__2]
        tex_3=texx[:indice_find__1]+"3F"+texx[indice_find__1:indice_find__2]+"3F"+texx[indice_find__2]
        tex_4=texx[:indice_find__1]+"4F"+texx[indice_find__1:indice_find__2]+"4F"+texx[indice_find__2]
        celda=hoja.cell(row=4,column=17)
        celda.value=texx
        celda=hoja.cell(row=5,column=17)
        celda.value=tex_2
        celda=hoja.cell(row=7,column=17)
        celda.value=tex_3
        celda=hoja.cell(row=8,column=17)
        celda.value=tex_4    
    ######################################!
 ###########################################SCRIPT  ANTES-NOC CABLE##########################################   

    texto="SCRIPT  ANTES-NOC CABLE"
    celda=hoja.cell(row=17,column=6)
    celda.value=texto
    cell = worksheet.cell(row=17, column=6)
    red=Color(rgb='FF0000')
    red_font = Font(color=red)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font=red_font 
    texto="config"
    celda=hoja.cell(row=19,column=6)
    celda.value=texto

    sep=0
    sep2=0
    text_script=""
    #Extrae de la columna 'Up',el texto para mirar si es necesario asignar un valor u otro
    slot_valor=df['Up']
    slot_index=slot_valor.index
    slot_list=slot_index.to_list()
    indice_slot=slot_list[0]
    texto_slot=df.loc[indice_slot,"Up"]
    #print(texto_slot)
    #Si el texto extraido anteriormente contiene "U"
    if "U" in texto_slot:
        #Si la longitud de la lista extraida es igual a 4, buscara el caracter "/", para asignar datos a solo 4 celdas y las demas vacias
        if len(slot_list)==4:
            sep=texto_slot.find("/")
            sep2=texto_slot.find("/",sep+1)
            text_script=texto_slot[:sep2+2]
            text_script_a=text_script
            text_script_b=text_script
            text_script_c=text_script
            text_script_d=text_script
            text_script_e=""
            text_script_f=""
            text_script_g=""
            text_script_h=""
            
        elif len(slot_list)>4:
            #Si en la posicion 5 de la lista esta vacio, buscara el caracter "/", para asignar datos a solo 5 celdas y las demas vacias
            if slot_list[5]==None:
                sep=texto_slot.find("/")
                sep2=texto_slot.find("/",sep+1)
                text_script=texto_slot[:sep2+2]
                text_script_a=text_script
                text_script_b=text_script
                text_script_c=text_script
                text_script_d=text_script
                text_script_e=text_script
                text_script_f=""
                text_script_g=""
                text_script_h=""
                #Si en la posicion 6 de la lista esta vacio, buscara el caracter "/", para asignar datos a solo 6 celdas y las demas vacias
            elif slot_index[6]==None:
                sep=texto_slot.find("/")
                sep2=texto_slot.find("/",sep+1)
                text_script=texto_slot[:sep2+2]
                text_script_a=text_script
                text_script_b=text_script
                text_script_c=text_script
                text_script_d=text_script
                text_script_e=text_script
                text_script_f=text_script
                text_script_g=""
                text_script_h=""
                #Si en la posicion 7 de la lista esta vacio, buscara el caracter "/", para asignar datos a solo 7 celdas y las demas vacias
            elif slot_index[7]==None:
                sep=texto_slot.find("/")
                sep2=texto_slot.find("/",sep+1)
                text_script=texto_slot[:sep2+2]
                text_script_a=text_script
                text_script_b=text_script
                text_script_c=text_script
                text_script_d=text_script
                text_script_e=text_script
                text_script_f=text_script
                text_script_g=text_script
                text_script_h=""
                #si tiene valores en los 8 datos necesarios, buscara el caracter "/", para asignar datos a todos los espacios
            else:
                sep=texto_slot.find("/")
                sep2=texto_slot.find("/",sep+1)
                text_script=texto_slot[:sep2+2]
                text_script_a=text_script
                text_script_b=text_script
                text_script_c=text_script
                text_script_d=text_script
                text_script_e=text_script
                text_script_f=text_script
                text_script_g=text_script
                text_script_h=text_script
        else:
                sep=texto_slot.find("/")
                sep2=texto_slot.find("/",sep+1)
                text_script=texto_slot[:sep2+2]
                text_script_a=text_script
                text_script_b=text_script
                text_script_c=text_script
                text_script_d=""
                text_script_e=""
                text_script_f=""
                text_script_g=""
                text_script_h=""
    else:
        print(f"len_slot_list==>{len(slot_list)}")
        #Si la longitud de la lista extraida es igual a 4, buscara el caracter "/", para asignar datos a solo 4 celdas y las demas vacias
        if len(slot_list)==4:
            a=df.loc[slot_list[0],"Up"]
            b=df.loc[slot_list[1],"Up"]
            c=df.loc[slot_list[2],"Up"]
            d=df.loc[slot_list[3],"Up"]
            #print(f"a==>{a}")
            #print(f"b==>{b}")
            #print(f"c==>{c}")
            #print(f"d==>{d}")
            sepa=a.find("/")
            sepa2=a.find("/",sepa+1)
            text_script_a=a[:sepa2]
            sepb=b.find("/")
            sepb2=b.find("/",sepb+1)
            text_script_b=b[:sepb2]
            sepc=c.find("/")
            sepc2=c.find("/",sepc+1)
            text_script_c=c[:sepc2]
            sepd=d.find("/")
            sepd2=d.find("/",sepd+1)
            text_script_d=d[:sepd2]
            text_script_e=""
            text_script_f=""
            text_script_g=""
            text_script_h=""

        elif len(slot_list)>4:
            #Si en la posicion 5 de la lista esta vacio, buscara el caracter "/", para asignar datos a solo 5 celdas y las demas vacias
            if slot_list[5]==None:
                a=df.loc[slot_list[0],"Up"]
                b=df.loc[slot_list[1],"Up"]
                c=df.loc[slot_list[2],"Up"]
                d=df.loc[slot_list[3],"Up"]
                e=df.loc[slot_list[4],"Up"]
                #print(f"a==>{a}")
                #print(f"b==>{b}")
                #print(f"c==>{c}")
                #print(f"d==>{d}")
                sepa=a.find("/")
                sepa2=a.find("/",sepa+1)
                text_script_a=a[:sepa2]
                sepb=b.find("/")
                sepb2=b.find("/",sepb+1)
                text_script_b=b[:sepb2]
                sepc=c.find("/")
                sepc2=c.find("/",sepc+1)
                text_script_c=c[:sepc2]
                sepd=d.find("/")
                sepd2=d.find("/",sepd+1)
                text_script_d=d[:sepd2]

                sepe=e.find("/")
                sepe2=e.find("/",sepe+1)
                text_script_e=e[:sepe2]
                text_script_f=" "
                text_script_g=" "                
                text_script_h=" "
                #Si en la posicion 6 de la lista esta vacio, buscara el caracter "/", para asignar datos a solo 6 celdas y las demas vacias
            elif slot_list[6]==None:
                a=df.loc[slot_list[0],"Up"]
                b=df.loc[slot_list[1],"Up"]
                c=df.loc[slot_list[2],"Up"]
                d=df.loc[slot_list[3],"Up"]
                e=df.loc[slot_list[4],"Up"]
                f=df.loc[slot_list[5],"Up"]
                #print(f"a==>{a}")
                #print(f"b==>{b}")
                #print(f"c==>{c}")
                #print(f"d==>{d}")
                sepa=a.find("/")
                sepa2=a.find("/",sepa+1)
                text_script_a=a[:sepa2]
                sepb=b.find("/")
                sepb2=b.find("/",sepb+1)
                text_script_b=b[:sepb2]
                sepc=c.find("/")
                sepc2=c.find("/",sepc+1)
                text_script_c=c[:sepc2]
                sepd=d.find("/")
                sepd2=d.find("/",sepd+1)
                text_script_d=d[:sepd2]

                sepe=e.find("/")
                sepe2=e.find("/",sepe+1)
                text_script_e=e[:sepe2]
                sepf=f.find("/")
                sepf2=f.find("/",sepf+1)
                text_script_f=f[:sepf2]
                text_script_g=" "                
                text_script_h=" "
                #Si en la posicion 7 de la lista esta vacio, buscara el caracter "/", para asignar datos a solo 7 celdas y las demas vacias
            elif slot_list[7]==None:
                a=df.loc[slot_list[0],"Up"]
                b=df.loc[slot_list[1],"Up"]
                c=df.loc[slot_list[2],"Up"]
                d=df.loc[slot_list[3],"Up"]
                e=df.loc[slot_list[4],"Up"]
                f=df.loc[slot_list[5],"Up"]
                g=df.loc[slot_list[6],"Up"]
                #print(f"a==>{a}")
                #print(f"b==>{b}")
                #print(f"c==>{c}")
                #print(f"d==>{d}")
                sepa=a.find("/")
                sepa2=a.find("/",sepa+1)
                text_script_a=a[:sepa2]
                sepb=b.find("/")
                sepb2=b.find("/",sepb+1)
                text_script_b=b[:sepb2]
                sepc=c.find("/")
                sepc2=c.find("/",sepc+1)
                text_script_c=c[:sepc2]
                sepd=d.find("/")
                sepd2=d.find("/",sepd+1)
                text_script_d=d[:sepd2]

                sepe=e.find("/")
                sepe2=e.find("/",sepe+1)
                text_script_e=e[:sepe2]
                sepf=f.find("/")
                sepf2=f.find("/",sepf+1)
                text_script_f=f[:sepf2]
                sepg=g.find("/")
                sepg2=g.find("/",sepg+1)
                text_script_g=g[:sepg2]              
                text_script_h=" "
            #si tiene valores en los 8 datos necesarios, buscara el caracter "/", para asignar datos a todos los espacios
            else:
                
                a=df.loc[slot_list[0],"Up"]
                b=df.loc[slot_list[1],"Up"]
                c=df.loc[slot_list[2],"Up"]
                d=df.loc[slot_list[3],"Up"]
                e=df.loc[slot_list[4],"Up"]
                f=df.loc[slot_list[5],"Up"]
                g=df.loc[slot_list[6],"Up"]
                h=df.loc[slot_list[7],"Up"]
                #print(f"a==>{a}")
                #print(f"b==>{b}")
                #print(f"c==>{c}")
                #print(f"d==>{d}")
                sepa=a.find("/")
                sepa2=a.find("/",sepa+1)
                text_script_a=a[:sepa2]
                sepb=b.find("/")
                sepb2=b.find("/",sepb+1)
                text_script_b=b[:sepb2]
                sepc=c.find("/")
                sepc2=c.find("/",sepc+1)
                text_script_c=c[:sepc2]
                sepd=d.find("/")
                sepd2=d.find("/",sepd+1)
                text_script_d=d[:sepd2]

                sepe=e.find("/")
                sepe2=e.find("/",sepe+1)
                text_script_e=e[:sepe2]
                sepf=f.find("/")
                sepf2=f.find("/",sepf+1)
                text_script_f=f[:sepf2]
                sepg=g.find("/")
                sepg2=g.find("/",sepg+1)
                text_script_g=g[:sepg2]
                seph=h.find("/")
                seph2=h.find("/",seph+1)
                text_script_h=h[:seph2]        

        else:
            a=df.loc[slot_list[0],"Up"]
            b=df.loc[slot_list[1],"Up"]
            c=df.loc[slot_list[2],"Up"]
            #print(f"a==>{a}")
            #print(f"b==>{b}")
            #print(f"c==>{c}")
            sepa=a.find("/")
            sepa2=a.find("/",sepa+1)
            text_script_a=a[:sepa2]
            sepb=b.find("/")
            sepb2=b.find("/",sepb+1)
            text_script_b=b[:sepb2]
            sepc=c.find("/")
            sepc2=c.find("/",sepc+1)
            text_script_c=c[:sepc2]
            text_script_d=""
            text_script_e=""
            text_script_f=""
            text_script_g=""
            text_script_h=""

    #dependiendo de la condicion que sea asigna un valor u otro a la celda
    texto="interface upstream "+text_script_a
    celda=hoja.cell(row=20,column=6)
    celda.value=texto
    texto='  description "PUERTO LIBRE"'
    celda=hoja.cell(row=21,column=6)
    celda.value=texto
    texto='  logical-channel 0 description "PUERTO LIBRE"'
    celda=hoja.cell(row=22,column=6)
    celda.value=texto
    texto="end"
    celda=hoja.cell(row=23,column=6)
    celda.value=texto
    texto="interface upstream "+text_script_b
    celda=hoja.cell(row=24,column=6)
    celda.value=texto
    texto='  description "PUERTO LIBRE"'
    celda=hoja.cell(row=25,column=6)
    celda.value=texto
    texto='  logical-channel 0 description "PUERTO LIBRE"'
    celda=hoja.cell(row=26,column=6)
    celda.value=texto
    texto="end"
    celda=hoja.cell(row=27,column=6)
    celda.value=texto
    texto="interface upstream "+text_script_c
    celda=hoja.cell(row=28,column=6)
    celda.value=texto
    texto='  description "PUERTO LIBRE"'
    celda=hoja.cell(row=29,column=6)
    celda.value=texto
    texto='  logical-channel 0 description "PUERTO LIBRE"'
    celda=hoja.cell(row=30,column=6)
    celda.value=texto
    texto="end"
    celda=hoja.cell(row=31,column=6)
    celda.value=texto
    texto="interface upstream "+text_script_d
    celda=hoja.cell(row=32,column=6)
    celda.value=texto
    texto='  description "PUERTO LIBRE"'
    celda=hoja.cell(row=33,column=6)
    celda.value=texto
    texto='  logical-channel 0 description "PUERTO LIBRE"'
    celda=hoja.cell(row=34,column=6)
    celda.value=texto  

    texto="end"
    celda=hoja.cell(row=35,column=6)
    celda.value=texto
    texto="interface upstream "+text_script_e
    celda=hoja.cell(row=36,column=6)
    celda.value=texto
    texto='  description "PUERTO LIBRE"'
    celda=hoja.cell(row=37,column=6)
    celda.value=texto
    texto='  logical-channel 0 description "PUERTO LIBRE"'
    celda=hoja.cell(row=38,column=6)
    celda.value=texto             
    texto="end"
    celda=hoja.cell(row=39,column=6)
    celda.value=texto
    texto="interface upstream "+text_script_f
    celda=hoja.cell(row=40,column=6)
    celda.value=texto
    texto='  description "PUERTO LIBRE"'
    celda=hoja.cell(row=41,column=6)
    celda.value=texto
    texto='  logical-channel 0 description "PUERTO LIBRE"'
    celda=hoja.cell(row=42,column=6)
    celda.value=texto 
    texto="end"
    celda=hoja.cell(row=43,column=6)
    celda.value=texto
    texto="interface upstream "+text_script_g
    celda=hoja.cell(row=44,column=6)
    celda.value=texto
    texto='  description "PUERTO LIBRE"'
    celda=hoja.cell(row=45,column=6)
    celda.value=texto
    texto='  logical-channel 0 description "PUERTO LIBRE"'
    celda=hoja.cell(row=46,column=6)
    celda.value=texto 
    texto="end"
    celda=hoja.cell(row=47,column=6)
    celda.value=texto
    texto="interface upstream "+text_script_h
    celda=hoja.cell(row=48,column=6)
    celda.value=texto
    texto='  description "PUERTO LIBRE"'
    celda=hoja.cell(row=49,column=6)
    celda.value=texto
    texto='  logical-channel 0 description "PUERTO LIBRE"'
    celda=hoja.cell(row=50,column=6)
    celda.value=texto 

    #Extrar el texto de la columna "Description"
    description_valor=df['Description']
    description_index=description_valor.index
    description_list=description_index.to_list()
    #print(f"slot_valor==>{description_valor}")
    #print(f"slot_list==>{description_list}")
    indice_description=description_list[0]
    text_description=df.loc[indice_description,"Description"]
    #Busca palabras clave para solo extraer el nombre del nodo, al cual se le esta generando el diseño
    sep=text_description.find("DO")
    sep2=text_description.find("(")
    text_script=text_description[sep+2:sep2]   

    texto="config"
    celda=hoja.cell(row=54,column=6)
    celda.value=texto
    texto="no service group "+text_script
    celda=hoja.cell(row=55,column=6)
    celda.value=texto
    texto="exit"
    celda=hoja.cell(row=56,column=6)
    celda.value=texto

    texto="NOTA: SI HAY CAMBIO DE IP"
    hoja.merge_cells(start_row=16, start_column=7, end_row=16, end_column=10)
    celda=hoja.cell(row=16,column=7)
    celda.value=texto
    cell_range_row16 = worksheet['G16:J16']
    cell = worksheet.cell(row=17, column=7)
    red_2=Color(rgb='FF0000')
    white=Color(rgb='FFFFFF')
    white_font = Font(color=white)
    relleno_2 = PatternFill(start_color=red_2, end_color=red_2, fill_type='solid')
    celda.alignment = Alignment(horizontal='center', vertical='center')
    celda.font=white_font
    for cells in cell_range_row16:
        for cell in cells:
            cell.fill=relleno_2                
    archivo_excel.save()    # Guardar archivo Excel
#funcion para realizar split de los elementos de la columna "Puerto DAAS"
def get_x(s, n=2):
    elements = s.split('/')
    if len(elements) >= n+1:
        return elements[-(n+1)]
    else:
        return None
