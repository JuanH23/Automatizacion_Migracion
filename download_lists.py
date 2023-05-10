from office365_api import SharePoint
import sys
import csv
from pathlib import PurePath
from pathlib import Path
from openpyxl import Workbook
import pandas as pd
import time
import threading
import os
import ssl
from openpyxl import load_workbook
from Estructura_principal_FINAL import *
from tqdm import tqdm
from PyQt5.QtWidgets import QMessageBox
def Type_file(file_name,export_type):
    #Dependiendo del tipo de archivo que se coloque va a a añadir la estensión
    #mas el nombre que le entra como parametro
    if export_type == 'Excel':
        file_name_export='.'.join([file_name,'xlsx']) 
    elif export_type == 'CSV':
        file_name_export='.'.join([file_name,'csv'])
    else:
        file_name_export=file_name
    return file_name_export

   
def download_list(list_name,export_type,dir_path,file_name,progress_callback=None):
    #Llama con un hilo la función save_Execel para ejecutar en segundo plano
    ssl._create_default_https_context=ssl._create_unverified_context 
    sp_list=SharePoint().get_list(list_name)
    total_items=len(sp_list)
    
    if export_type == 'Excel':
                
        file=threading.Thread(target=save_Execel, args=(sp_list, dir_path, file_name,progress_callback))
        file.start()
            
    elif export_type =='CSV':
        save_file_csv(sp_list,dir_path,file_name)
    else:
        print("No se puede Descargar ese tipo de archivo")
    

def save_file_csv(list_items,dir_path,file_name):

    dir_file_path=PurePath(dir_path,file_name)
    with open (dir_file_path,'w',newline='\n',encoding='utf8') as f:
        header=list_items[0].properties.keys()
        w=csv.DictWriter(f,header)
        w.writeheader()
        for item in list_items:
            w.writerow(item.properties)


def save_Execel(list_items,dir_path,file_name,progress_callback=None):
    ssl._create_default_https_context=ssl._create_unverified_context   
    dir_file_path=Path(dir_path,file_name).with_suffix('.xlsx') 
    # dir_file_path=PurePath(dir_path,file_name)
    total_items = len(list_items)
    wb= Workbook()
    ws=wb.active   
    #Obtiene las cabeceras de la lista
    header=list_items[0].properties.keys()
    
    #Escribe las columnas en la primera fila
    for idx,name in enumerate(header):
        ws.cell(row=1, column=idx+1,value=name)
    #Comienza a escribir los items desde la segunda fila
    row=2
    with tqdm(total=total_items, desc='Descargando archivo') as pbar:
        for i,dict_obj in enumerate(list_items, start=1):
                
                    
                    for idx, item in enumerate(dict_obj.properties.items()):
                        ws.cell(row=row,column=idx+1,value=item[1])
                        if progress_callback:
                            
                            progress_callback(pbar, row, total_items)                        
                    row+=1
                    

   
    print("AAAAAA")
    dir_path=Path(dir_path)
    dir_path.mkdir(parents=True,exist_ok=True)
    wb.save(dir_file_path)
    

    df=read_excel_to_dataframe(dir_file_path,file_name)
    df.to_excel(dir_file_path,index=False)
    print(f"df==>{df}")
    
    
def read_excel_to_dataframe(file_path,file_name):

    wb = load_workbook(file_path)
    
    print(f"file_name==>{file_name}")
    ws = wb.active
    data = ws.values
    headers = next(data)
    # Crear el DataFrame con la data y las cabeceras de columna
    df = pd.DataFrame(data, columns=headers)
    #Revisa el nombre del archivo que este, y mira en el las cabeceras, 
    # si coincide uno u otro filtra el archivo, para solo obtener lo necesario
    if "Ocupacion-Harmonic_COS" in file_name:#!COS_TERMINAR CUANDO SE SEPAREN LOS ARCHIVOS EN LA DESCARGA
        cont1=0
        print("a")
        cabeceras=list(df.columns)
        headers=['IP','Dispositivo','Puerto','ID','moka','status','ptp']
        for header in headers:
            if header in cabeceras:
                cont1+=1
                print(f"cont1==>{cont1}")
                if cont1==4:
                    df=df.loc[:,['IP','Dispositivo','Puerto','ID','moka','status','ptp']]
    elif "Arris_SCMSummary" in file_name:
        cont2=0
        print("b")
        cabeceras=list(df.columns)
        headers=['CMTS','Up','Mac','Conn','Total','Oper','Disable','Init','Offline','Description']
        for header in headers:
            if header in cabeceras:
                cont2+=1
                print(f"cont1==>{cont2}")
                if cont2==4:
                    df=df.loc[:,['CMTS','Up','Mac','Conn','Total','Oper','Disable','Init','Offline','Description']] 
    elif "Casa_SCMSummary" in file_name:
        cont3=0
        print("c")
        cabeceras=list(df.columns)
        headers=['CMTS','Upstream','Total','Active','Registered','Secondary','offline','Bonding','NonBonding','Description']
        for header in headers:
            if header in cabeceras:
                cont3+=1
                print(f"cont1==>{cont3}")
                if cont3==4:
                    df=df.loc[:,['CMTS','Upstream','Total','Active','Registered','Secondary','offline','Bonding','NonBonding','Description']]
    elif "Ocupacion- RPHY Harmonic_DAAS" in file_name:##!DAAS
        cont4=0
        print("d")
        cabeceras=list(df.columns)
        headers=['IP','Dispositivo','Puerto','status','status_2','ptp']
        for header in headers:
            if header in cabeceras:
                cont4+=1
                print(f"cont1==>{cont4}")
                if cont4==4:
                    df=df.loc[:,['IP','Dispositivo','Puerto','status','stat2','ptp']]                          
    return df

if __name__ =='__main__':
    pass
